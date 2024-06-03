from aiogram import Dispatcher, types
from aiogram.types import InputFile
from aiogram.dispatcher import FSMContext
from tg_bot.DBSM import all_link, all_user, delete_users_range, delete_link, add_link, all_time_loop, delete_loop_range
from tg_bot.keyboards import users_kb, links_kb, loop_kb
from openpyxl import Workbook
from openpyxl.styles import Font
from tg_bot.states import admin
import os


def register_handlers(dp: Dispatcher) -> None:
    dp.register_message_handler(delete_user_proc, state = admin.del_user)
    dp.register_message_handler(del_link, state = admin.del_link)
    dp.register_message_handler(add_link_proc, state = admin.add_link)
    dp.register_message_handler(loop_delete_process, state=admin.del_loop)
    dp.register_message_handler(links_admin, commands=['links'], state = None)
    dp.register_message_handler(users_admin, commands=['users'], state= None)
    dp.register_message_handler(loop_admin, commands=['reviews'], state= None)
    dp.register_callback_query_handler(loop_delete, text="loop_delete")
    dp.register_callback_query_handler(delete_user, text="user_delete", state = None)
    dp.register_callback_query_handler(links_process, text_startswith = "link", state = None)

async def users_admin(message: types.Message, state: FSMContext) -> None:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    sheet = wb.create_sheet("Пользователи", 0)
    res = all_user()
    sheet["A1"] = "id"
    sheet['A1'].font = Font(color="FF0000")  
    sheet["B1"] = "Юзернейм"
    sheet['B1'].font = Font(color="FF0000")  
    sheet["C1"] = "Шаг"
    sheet['C1'].font = Font(color="FF0000")  

    for i in range(len(res)):
        sheet[f"A{i+2}"] = res[i]["id"]
        sheet[f"B{i+2}"] = f"https://vk.com/{res[i]['user']}"
        sheet[f"C{i+2}"] = res[i]["step"]

    if not os.path.exists("tg_bot/excel"):
        os.mkdir('tg_bot/excel')
    wb.save("tg_bot/excel/Пользователи.xlsx")
    await message.answer_document(document= InputFile("tg_bot/excel/Пользователи.xlsx"), reply_markup=users_kb())
    os.remove("tg_bot/excel/Пользователи.xlsx")
    


async def delete_user(call: types.CallbackQuery, state: FSMContext) -> None:
    await call.message.answer("Ввведите диапазон удаления пользователей в статистике через пробел без букв и сторонних символов.\n Формат типа: 1 5")
    await admin.del_user.set()


async def loop_admin(message: types.Message, state: FSMContext):
    wb = Workbook()
    wb.remove(wb["Sheet"])
    data = all_time_loop()
    sheet = wb.create_sheet("Пользователи", 0)
    sheet["A1"] = "id"
    sheet['A1'].font = Font(color="FF0000")  
    sheet["B1"] = "Юзернейм"
    sheet['B1'].font = Font(color="FF0000")
    sheet["C1"] = "Шаг 2"
    sheet['C1'].font = Font(color="FF0000")
    sheet["D1"] = "Шаг 3"
    sheet['D1'].font = Font(color="FF0000")
    sheet["E1"] = "Шаг 4"
    sheet['E1'].font = Font(color="FF0000")
    sheet["F1"] = "Шаг 5"
    sheet['F1'].font = Font(color="FF0000")
    sheet["G1"] = "Шаг 6"
    sheet['G1'].font = Font(color="FF0000")
    sheet["H1"] = "Шаг 7"
    sheet['H1'].font = Font(color="FF0000")
    sheet["I1"] = "Шаг 10"
    sheet['I1'].font = Font(color="FF0000")
    for i in range(len(data)):
        sheet[f"A{i+2}"] = data[i]["id"]
        sheet[f"B{i+2}"] = f"https://vk.com/{data[i]['user']}"
        sheet[f"C{i+2}"] = data[i]["step2"]
        sheet[f"D{i+2}"] = data[i]["step3"]
        sheet[f"E{i+2}"] = data[i]["step4"]
        sheet[f"F{i+2}"] = data[i]["step5"]
        sheet[f"G{i+2}"] = data[i]["step6"]
        sheet[f"H{i+2}"] = data[i]["step7"]
        sheet[f"I{i+2}"] = data[i]["step10"]
    if not os.path.exists("tg_bot/excel"):
        os.mkdir('tg_bot/excel')
    wb.save("tg_bot/excel/Пользователи по времени.xlsx")
    await message.answer_document(document= InputFile("tg_bot/excel/Пользователи по времени.xlsx"), reply_markup=loop_kb())
    os.remove("tg_bot/excel/Пользователи по времени.xlsx")

async def loop_delete(call: types.CallbackQuery, state: FSMContext):
    await call.message.answer("Ввведите диапазон удаления пользователей в статистике по времени через пробел без букв и сторонних символов.\n Формат типа: 1 5")
    await admin.del_loop.set()

async def loop_delete_process(message: types.Message, state: FSMContext):
    data = message.text.split()
    delete_loop_range(int(data[0]), int(data[1]))
    await message.answer("Пользователи в статистике по времени удалены")
    await state.finish()


async def delete_user_proc(message: types.Message, state: FSMContext) -> None:
    data = message.text.split()
    delete_users_range(int(data[0]), int(data[1]))
    await message.answer("Пользователи в статистике удалены")
    await state.finish()


async def links_admin(message: types.Message, state: FSMContext) -> None:
    text = ""
    data = all_link()
    for i in data:
        text += f"Ссылка {i['id']}: {i['link']}\n"
    await message.answer(text, reply_markup= links_kb())


async def links_process(call: types.CallbackQuery, state: FSMContext):
    data = call.data.split('_')[1]
    if data == "add":
        await call.message.answer("Введите ссылку, которую хотите добавить, в следующем сообщении")
        await admin.add_link.set()
    else:
        await call.message.answer("Введите номер ссылки, которую хотите убрать, в следующем сообщении цифрой без букв и сторонних символов")
        await admin.del_link.set()


async def del_link(message: types.Message, state: FSMContext):
    delete_link(message.text)
    await message.answer("Ссылка удалена")
    await state.finish()


async def add_link_proc(message: types.Message, state: FSMContext):
    add_link(message.text)
    await message.answer("Ссылка добавлена")
    await state.finish()




